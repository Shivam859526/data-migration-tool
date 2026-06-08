"""Migration report generation — JSON and Excel output."""

import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from config.constants import REPORT_FOLDER
from utils.file_manager import FileManager
from utils.logger import Logger

logger = Logger.get_logger("report")


class ReportEngine:
    """Generates migration summary reports in JSON and Excel formats."""

    def __init__(self) -> None:
        FileManager.create_directories()
        self._report: Dict[str, Any] = {
            "start_time": None,
            "end_time": None,
            "duration_seconds": 0,
            "job_id": None,
            "tables_migrated": [],
            "total_rows_migrated": 0,
            "failures": [],
            "validation_results": [],
            "status": "PENDING",
        }

    def start(self, job_id: str) -> None:
        self._report["job_id"] = job_id
        self._report["start_time"] = datetime.now().isoformat()
        self._report["status"] = "RUNNING"

    def record_table(
        self,
        table_name: str,
        rows_migrated: int,
        status: str = "SUCCESS",
        error: Optional[str] = None,
    ) -> None:
        entry = {
            "table": table_name,
            "rows_migrated": rows_migrated,
            "status": status,
        }
        if error:
            entry["error"] = error
            self._report["failures"].append(entry)
        self._report["tables_migrated"].append(entry)
        self._report["total_rows_migrated"] += rows_migrated

    def record_validation(self, validation_result: Dict[str, Any]) -> None:
        self._report["validation_results"].append(validation_result)

    def finish(self, status: str = "COMPLETED") -> Dict[str, Any]:
        end = datetime.now()
        self._report["end_time"] = end.isoformat()
        if self._report["start_time"]:
            start = datetime.fromisoformat(self._report["start_time"])
            self._report["duration_seconds"] = (end - start).total_seconds()
        self._report["status"] = status
        return self._report

    def get_report(self) -> Dict[str, Any]:
        return self._report

    def export_json(self, filename: Optional[str] = None) -> str:
        if not filename:
            job_id = self._report.get("job_id", "unknown")
            filename = f"migration_report_{job_id}.json"
        path = os.path.join(REPORT_FOLDER, filename)
        with open(path, "w", encoding="utf-8") as file:
            json.dump(self._report, file, indent=4, default=str)
        logger.info("JSON report exported to %s", path)
        return path

    def export_excel(self, filename: Optional[str] = None) -> str:
        if not filename:
            job_id = self._report.get("job_id", "unknown")
            filename = f"migration_report_{job_id}.xlsx"

        path = os.path.join(REPORT_FOLDER, filename)
        wb = Workbook()

        summary = wb.active
        summary.title = "Summary"
        summary.append(["Migration Report"])
        summary["A1"].font = Font(bold=True, size=14)

        summary_fields = [
            ("Job ID", self._report.get("job_id")),
            ("Status", self._report.get("status")),
            ("Start Time", self._report.get("start_time")),
            ("End Time", self._report.get("end_time")),
            ("Duration (seconds)", self._report.get("duration_seconds")),
            ("Total Rows Migrated", self._report.get("total_rows_migrated")),
            ("Failures", len(self._report.get("failures", []))),
        ]
        for label, value in summary_fields:
            summary.append([label, value])

        tables_sheet = wb.create_sheet("Tables")
        tables_sheet.append(["Table", "Rows Migrated", "Status", "Error"])
        for entry in self._report.get("tables_migrated", []):
            tables_sheet.append([
                entry.get("table"),
                entry.get("rows_migrated"),
                entry.get("status"),
                entry.get("error", ""),
            ])

        validation_sheet = wb.create_sheet("Validation")
        validation_sheet.append([
            "Table", "Source Rows", "Target Rows", "Status",
        ])
        for vr in self._report.get("validation_results", []):
            validation_sheet.append([
                vr.get("table"),
                vr.get("source_rows"),
                vr.get("target_rows"),
                vr.get("status"),
            ])

        failures_sheet = wb.create_sheet("Failures")
        failures_sheet.append(["Table", "Rows", "Status", "Error"])
        for failure in self._report.get("failures", []):
            failures_sheet.append([
                failure.get("table"),
                failure.get("rows_migrated"),
                failure.get("status"),
                failure.get("error", ""),
            ])

        wb.save(path)
        logger.info("Excel report exported to %s", path)
        return path

    def export_all(self) -> Dict[str, str]:
        return {
            "json": self.export_json(),
            "excel": self.export_excel(),
        }
